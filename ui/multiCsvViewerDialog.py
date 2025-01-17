import os
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PySide6.QtWidgets import QApplication, QDialog, QVBoxLayout, QTabWidget, QTableWidget, QTableWidgetItem, QMessageBox

class MultiCsvViewerDialog(QDialog):
    def __init__(self, folder_path, output_csv_filename="all_info.xlsx"):
        super().__init__()
        self.setWindowTitle("CSV Viewer")
        self.resize(800, 600)

        # Layout to hold the tab widget
        layout = QVBoxLayout()

        # Tab widget to switch between CSV files
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)

        # Load CSV files from the provided folder path
        self.load_folder(folder_path)

        # If an output path is provided, merge all CSV files into one
        if output_csv_filename:
            self.merge_csv_to_excel(folder_path, output_csv_filename)

        self.setLayout(layout)

    def load_folder(self, folder_path):
        # List all CSV files in the provided folder path
        if not os.path.exists(folder_path):
            print(f"Folder {folder_path} does not exist.")
            return

        csv_files = [f for f in os.listdir(folder_path) if f.endswith('.csv')]

        if not csv_files:
            print("No CSV files found in the folder.")
            return

        # Clear previous tabs if any
        self.tabs.clear()

        # Create a new tab for each CSV file
        for csv_file in csv_files:
            csv_path = os.path.join(folder_path, csv_file)
            self.add_csv_tab(csv_path, csv_file)

    def add_csv_tab(self, csv_path, tab_name):
        # Create a new table widget
        table_widget = QTableWidget()

        # Load the CSV and populate the table
        self.load_csv_to_table(csv_path, table_widget)

        # Add the table widget to the tab widget
        self.tabs.addTab(table_widget, tab_name)

    def load_csv_to_table(self, csv_path, table_widget):
        try:
            with open(csv_path, 'r') as csvfile:
                reader = csv.reader(csvfile)
                rows = list(reader)

                # Set the number of rows and columns based on the CSV data
                if rows:
                    table_widget.setRowCount(len(rows))
                    table_widget.setColumnCount(len(rows[0]))

                    # Initialize a list to track the maximum length of content in each column
                    max_column_width = [0] * len(rows[0])

                    # Fill the table with CSV data
                    for row_idx, row in enumerate(rows):
                        for col_idx, cell in enumerate(row):
                            table_widget.setItem(row_idx, col_idx, QTableWidgetItem(cell))

                            # Update the maximum width of the column based on the current cell's content
                            max_column_width[col_idx] = max(max_column_width[col_idx], len(cell))

                    # Set the column widths based on the maximum content length
                    for col_idx, width in enumerate(max_column_width):
                        table_widget.setColumnWidth(col_idx, width * 15)  # Adjust the factor as needed

        except Exception as e:
            print(f"Error loading CSV file {csv_path}: {e}")

    def merge_csv_to_excel(self, folder_path, output_csv_filename):

        output_csv_path = os.path.join(folder_path, output_csv_filename)
        # Check if the output file already exists
        if os.path.exists(output_csv_path):
            # File already exists, show a confirmation dialog
            reply = QMessageBox.question(self, "File Exists", 
                                        f"The file {output_csv_path} already exists. Do you want to overwrite it?",
                                        QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

            if reply == QMessageBox.No:
                return

        # Create a new workbook
        wb = Workbook()

        # List all CSV files in the folder
        csv_files = [f for f in os.listdir(folder_path) if f.endswith('.csv')]

        if not csv_files:
            print("No CSV files found in the folder to merge.")
            return

        for csv_file in csv_files:
            csv_path = os.path.join(folder_path, csv_file)
            try:
                # Create a new worksheet for each CSV file
                sheet_name = os.path.splitext(csv_file)[0]  # Use the filename as the sheet name
                sheet = wb.create_sheet(title=sheet_name)

                # Read the CSV file and add its content to the worksheet
                with open(csv_path, 'r') as infile:
                    reader = csv.reader(infile)
                    max_column_width = []

                    for row_idx, row in enumerate(reader):
                        # Write each row of data to the worksheet
                        for col_idx, cell in enumerate(row):
                            sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell)
                            
                            # Update the maximum column width based on cell content length
                            if len(max_column_width) <= col_idx:
                                max_column_width.append(len(cell))  # Initialize with the current cell length
                            else:
                                max_column_width[col_idx] = max(max_column_width[col_idx], len(cell))

                    # Set column widths based on the maximum content length
                    for col_idx, max_len in enumerate(max_column_width):
                        # Set column width, adding some padding to ensure content isn't cut off
                        sheet.column_dimensions[get_column_letter(col_idx + 1)].width = max_len * 2

            except Exception as e:
                print(f"Error reading {csv_file}: {e}")


        # Remove the default sheet created by openpyxl
        del wb['Sheet']    
            
        # Save the workbook as an Excel file
        try:
            wb.save(output_csv_path)
            print(f"CSV files merged successfully into {output_csv_path}")
        except Exception as e:
            print(f"Error saving Excel file: {e}")




